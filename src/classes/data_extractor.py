# Classe responsável pela extração e organização dos dados da página de resultados

from bs4 import BeautifulSoup
import pandas as pd
import os
import sys
import platform
from datetime import datetime
from src.classes.file.path_manager import obter_caminho_dados


class DataExtractor:
    # Classe responsável por extrair dados da página de resultados e salvar em Excel
    
    def __init__(self, diretorio_base="arquivos_baixados"):
        # Inicializa o extrator de dados
        # Usa caminho compatível com executável
        self.diretorio_base = obter_caminho_dados(diretorio_base)
        self.data_hoje = datetime.now().strftime("%Y-%m-%d")
        self.diretorio_saida = os.path.join(self.diretorio_base, self.data_hoje)
        self.dados_extraidos = []
        
        # Cria o diretório de saída se não existir
        self._criar_diretorio_saida()
    
    def _criar_diretorio_saida(self):
        # Cria a estrutura de diretórios necessária:
        try:
            # Cria o diretório base se não existir
            if not os.path.exists(self.diretorio_base):
                os.makedirs(self.diretorio_base)
            
            # Cria a subpasta da data atual se não existir
            if not os.path.exists(self.diretorio_saida):
                os.makedirs(self.diretorio_saida)
                
        except Exception:
            pass
    
    def extrair_html_pagina(self, navegador):
        # Extrai o HTML completo da página atual usando Selenium
        try:
            # Obtém o HTML completo da página atual
            html_completo = navegador.page_source
            
            if html_completo:
                return html_completo
            else:
                return None
                
        except Exception:
            return None
    
    def analisar_estrutura_tabela(self, html):
        # Analisa a estrutura HTML para encontrar a tabela de dados
        try:
            # Cria objeto BeautifulSoup para análise do HTML
            soup = BeautifulSoup(html, 'html.parser')
            
            # Procura por diferentes tipos de tabelas que podem conter os dados
            tabelas_encontradas = []
            
            # Estratégia 1: Tabelas HTML tradicionais
            tabelas_html = soup.find_all('table')
            tabelas_encontradas.extend(tabelas_html)
            
            # Estratégia 2: Divs com classes relacionadas a tabelas/dados
            divs_tabela = soup.find_all('div', class_=lambda x: x and any(
                palavra in str(x).lower() for palavra in ['table', 'grid', 'data', 'row', 'list']
            ))
            
            # Estratégia 3: Procurar por texto específico que indica dados
            elementos_com_data = soup.find_all(text=lambda text: text and any(
                palavra in str(text).lower() for palavra in ['data', 'parcela', 'valor', 'distribuído']
            ))
            
            if tabelas_html or divs_tabela or elementos_com_data:
                return soup
            else:
                return None
                
        except Exception:
            return None
    
    def extrair_dados_tabela(self, soup):
        # Extrai os dados específicos (DATA, PARCELA, VALOR DISTRIBUÍDO) da tabela
        try:
            dados = []
            
            # Estratégia 1: Tabelas HTML tradicionais
            tabelas = soup.find_all('table')
            for tabela in tabelas:
                dados_tabela = self._extrair_de_tabela_html(tabela)
                dados.extend(dados_tabela)
            
            # Estratégia 2: Estruturas baseadas em divs
            if not dados:
                dados = self._extrair_de_divs(soup)
            
            # Estratégia 3: Busca por padrões de texto
            if not dados:
                dados = self._extrair_por_padroes_texto(soup)
            
            return dados
            
        except Exception:
            return []
    
    def _extrair_de_tabela_html(self, tabela):
        # Extrai dados de uma tabela HTML tradicional
        dados = []
        
        try:
            # Procura linhas da tabela
            linhas = tabela.find_all('tr')
            
            # Processa TODAS as linhas, sem pular nenhuma
            for i, linha in enumerate(linhas):
                colunas = linha.find_all(['td', 'th'])
                
                # Extrai texto de cada coluna, preenchendo com vazio se não existir
                data_col = colunas[0].get_text(strip=True) if len(colunas) > 0 else ''
                parcela_col = colunas[1].get_text(strip=True) if len(colunas) > 1 else ''
                valor_col = colunas[2].get_text(strip=True) if len(colunas) > 2 else ''
                
                registro = {
                    'data': data_col,
                    'parcela': parcela_col,
                    'valor_distribuido': valor_col
                }
                
                # Adiciona TODAS as linhas, incluindo vazias e títulos
                dados.append(registro)
            
        except Exception:
            pass
        
        return dados
    
    def _extrair_de_divs(self, soup):
        # Extrai dados de estruturas baseadas em divs
        dados = []
        
        try:
            # Procura por divs que podem conter linhas de dados
            possíveis_linhas = soup.find_all('div', class_=lambda x: x and any(
                palavra in str(x).lower() for palavra in ['row', 'item', 'line', 'record']
            ))
            
            for linha in possíveis_linhas:
                # Procura por texto que contenha padrões de data, parcela e valor
                texto_linha = linha.get_text(strip=True)
                
                # Verifica se contém elementos que parecem dados financeiros
                if self._parece_linha_dados(texto_linha):
                    # Tenta extrair os componentes da linha
                    componentes = self._extrair_componentes_linha(texto_linha)
                    if componentes:
                        dados.append(componentes)
        
        except Exception:
            pass
        
        return dados
    
    def _extrair_por_padroes_texto(self, soup):
        # Extrai dados procurando por padrões específicos no texto
        dados = []
        
        try:
            # Procura por todo o texto da página
            texto_completo = soup.get_text()
            
            # Divide em linhas e procura por padrões
            linhas = texto_completo.split('\n')
            
            for linha in linhas:
                linha = linha.strip()
                if self._parece_linha_dados(linha):
                    componentes = self._extrair_componentes_linha(linha)
                    if componentes:
                        dados.append(componentes)
        
        except Exception:
            pass
        
        return dados
    
    def _parece_linha_dados(self, texto):
        # Verifica se uma linha de texto parece conter dados financeiros
        if not texto or len(texto) < 10:
            return False
        
        # Procura por padrões que indicam dados financeiros
        indicadores = [
            # Padrões de data (DD.MM.AAAA, DD/MM/AAAA)
            any(char.isdigit() for char in texto[:10]),
            # Palavras-chave relacionadas a parcelas
            any(palavra in texto.lower() for palavra in ['parcela', 'ipi', 'fpm', 'distribuição']),
            # Padrões de valor monetário
            any(char in texto for char in [',', 'R$', '.']),
        ]
        
        # Considera linha de dados se tiver pelo menos 2 indicadores
        return sum(indicadores) >= 2
    
    def _extrair_componentes_linha(self, linha):
        # Extrai os componentes (data, parcela, valor) de uma linha de texto
        try:
            import re
            
            # Remove espaços extras e caracteres especiais
            linha_limpa = ' '.join(linha.split())
            
            # Padrões regex para identificar componentes
            padrao_data = r'\b\d{1,2}[./]\d{1,2}[./]\d{4}\b'
            padrao_valor = r'\b\d{1,3}(?:[.,]\d{3})*[.,]\d{2}\b'
            
            # Procura data
            match_data = re.search(padrao_data, linha_limpa)
            data = match_data.group() if match_data else ''
            
            # Procura valor
            match_valor = re.search(padrao_valor, linha_limpa)
            valor = match_valor.group() if match_valor else ''
            
            # O que sobra é considerado parcela
            parcela = linha_limpa
            if match_data:
                parcela = parcela.replace(match_data.group(), '')
            if match_valor:
                parcela = parcela.replace(match_valor.group(), '')
            parcela = ' '.join(parcela.split())  # Remove espaços extras
            
            # Só retorna se tiver pelo menos data ou valor
            if data or valor:
                return {
                    'data': data,
                    'parcela': parcela,
                    'valor_distribuido': valor
                }
            
            return None
            
        except Exception:
            return None
    
    def salvar_dados_excel(self, dados, cidade, data_consulta=None):
        # Salva os dados extraídos em um arquivo Excel organizado
        try:
            if not dados:
                return None
            
            # Cria DataFrame apenas com as 3 colunas principais
            df = pd.DataFrame(dados)
            
            # Mantém apenas as colunas necessárias
            colunas_necessarias = ['data', 'parcela', 'valor_distribuido']
            df_final = df[colunas_necessarias].copy()
            
            # Renomeia as colunas para o formato do site
            df_final.columns = ['DATA', 'PARCELA', 'VALOR DISTRIBUÍDO (R$)']
            
            # Gera nome do arquivo: cidade_HHMMSS.xlsx
            timestamp = datetime.now().strftime("%H%M%S")
            cidade_formatada = cidade.replace(' ', '_').replace('/', '_').replace('\\', '_')
            nome_arquivo = f"{cidade_formatada}_{timestamp}.xlsx"
            caminho_arquivo = os.path.join(self.diretorio_saida, nome_arquivo)
            
            # Salva no Excel com formatação personalizada
            self._salvar_excel_formatado(df_final, caminho_arquivo, cidade)
            
            return caminho_arquivo
            
        except Exception:
            return None
    
    def _salvar_excel_formatado(self, df, caminho_arquivo, cidade):
        # Salva o Excel com formatação personalizada incluindo cores e layout
        try:
            from openpyxl import Workbook
            from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
            from openpyxl.utils.dataframe import dataframe_to_rows
            
            # Cria workbook
            wb = Workbook()
            ws = wb.active
            ws.title = "FPM - Demonstrativo"
            
            # Define estilos
            fonte_titulo = Font(name='Arial', size=14, bold=True, color='000000')
            fonte_subtitulo = Font(name='Arial', size=12, bold=True, color='000000')
            fonte_cabecalho = Font(name='Arial', size=11, bold=True, color='000000')
            fonte_dados = Font(name='Arial', size=10, color='000000')
            fonte_azul = Font(name='Arial', size=10, color='0000FF', bold=True)  # Azul
            fonte_vermelha = Font(name='Arial', size=10, color='FF0000', bold=True)  # Vermelho
            
            alinhamento_centro = Alignment(horizontal='center', vertical='center')
            alinhamento_esquerda = Alignment(horizontal='left', vertical='center')
            alinhamento_direita = Alignment(horizontal='right', vertical='center')
            
            borda_fina = Border(
                left=Side(style='thin'),
                right=Side(style='thin'),
                top=Side(style='thin'),
                bottom=Side(style='thin')
            )
            
            # Linha 1: Nome do município
            ws['A1'] = cidade.upper()
            ws['A1'].font = fonte_titulo
            ws['A1'].alignment = alinhamento_centro
            ws.merge_cells('A1:C1')
            
            # Linha 2: Título principal
            ws['A2'] = "FPM - FUNDO DE PARTICIPACAO DOS MUNICIPIOS"
            ws['A2'].font = fonte_subtitulo
            ws['A2'].alignment = alinhamento_centro
            ws.merge_cells('A2:C2')
            
            # Linha 3: Vazia para espaçamento
            
            # Linha 4: Cabeçalhos das colunas
            linha_cabecalho = 4
            colunas = ['DATA', 'PARCELA', 'VALOR DISTRIBUÍDO\n(R$)']
            
            for col_idx, coluna in enumerate(colunas, 1):
                cell = ws.cell(row=linha_cabecalho, column=col_idx, value=coluna)
                cell.font = fonte_cabecalho
                cell.alignment = alinhamento_centro
                cell.border = borda_fina
            
            # Adiciona os dados preservando todas as linhas originais
            linha_atual = linha_cabecalho + 1
            
            for idx, (_, row) in enumerate(df.iterrows()):
                data_atual = str(row['DATA'])
                
                # Coluna DATA
                ws.cell(row=linha_atual, column=1, value=data_atual).font = fonte_dados
                ws.cell(row=linha_atual, column=1).alignment = alinhamento_centro
                ws.cell(row=linha_atual, column=1).border = borda_fina
                
                # Coluna PARCELA
                ws.cell(row=linha_atual, column=2, value=row['PARCELA']).font = fonte_dados
                ws.cell(row=linha_atual, column=2).alignment = alinhamento_esquerda
                ws.cell(row=linha_atual, column=2).border = borda_fina
                
                # Coluna VALOR DISTRIBUÍDO com cores
                valor = str(row['VALOR DISTRIBUÍDO (R$)'])
                cell_valor = ws.cell(row=linha_atual, column=3, value=valor)
                cell_valor.alignment = alinhamento_direita
                cell_valor.border = borda_fina
                
                # Aplica cor baseada no valor (C = azul, D = vermelho)
                if valor.endswith('C'):
                    cell_valor.font = fonte_azul
                elif valor.endswith('D'):
                    cell_valor.font = fonte_vermelha
                else:
                    cell_valor.font = fonte_dados
                
                
                linha_atual += 1
            
            # Ajusta largura das colunas
            ws.column_dimensions['A'].width = 15  # DATA
            ws.column_dimensions['B'].width = 25  # PARCELA
            ws.column_dimensions['C'].width = 20  # VALOR DISTRIBUÍDO
            
            # Ajusta altura das linhas do cabeçalho
            ws.row_dimensions[1].height = 20  # Município
            ws.row_dimensions[2].height = 18  # Título
            ws.row_dimensions[4].height = 30  # Cabeçalhos
            
            # Salva o arquivo
            wb.save(caminho_arquivo)
            
        except Exception:
            # Fallback: salva sem formatação se houver erro
            df.to_excel(caminho_arquivo, index=False, engine='openpyxl')
    
    def processar_pagina_resultados(self, navegador, cidade):
        # Processo completo: extrai HTML → analisa → extrai dados → salva Excel
        try:
            # Passo 1: Extrair HTML da página
            html = self.extrair_html_pagina(navegador)
            if not html:
                return {'sucesso': False, 'erro': 'Falha ao extrair HTML'}
            
            # Passo 2: Analisar estrutura da tabela
            soup = self.analisar_estrutura_tabela(html)
            if not soup:
                return {'sucesso': False, 'erro': 'Falha ao analisar estrutura'}
            
            # Passo 3: Extrair dados da tabela
            dados = self.extrair_dados_tabela(soup)
            if not dados:
                return {'sucesso': False, 'erro': 'Nenhum dado encontrado na tabela'}
            
            # Passo 4: Salvar dados em Excel
            arquivo_salvo = self.salvar_dados_excel(dados, cidade)
            if not arquivo_salvo:
                return {'sucesso': False, 'erro': 'Falha ao salvar Excel'}
            
            # Retorna resultado de sucesso
            return {
                'sucesso': True,
                'registros_encontrados': len(dados),
                'arquivo_salvo': arquivo_salvo,
                'cidade': cidade
            }
            
        except Exception as e:
            return {'sucesso': False, 'erro': f'Erro inesperado: {e}'}
    