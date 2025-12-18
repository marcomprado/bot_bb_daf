"""
PDF to Excel Converter - AI-powered extraction.
This module processes PDF files containing government resolutions and generates a
consolidated Excel file with structured data extracted using AI.
"""

import sys
import os
from pathlib import Path

# Path adjustment for imports
sys.path.append(os.path.dirname(os.path.dirname(os.path.dirname(os.path.abspath(__file__)))))

# Standard library imports
import json
import re
import time
from datetime import datetime
from typing import Dict, List, Optional, Any

# Third-party imports
from dotenv import load_dotenv
import pymupdf4llm
from openai import OpenAI
import openai
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

# ============================================================================
# AI CONFIGURATION - Customize provider and model here
# ============================================================================

# AI Provider Settings
AI_PROVIDER = 'google-ai-studio'  # Options: 'openai', 'openrouter', 'google-ai-studio', 'nvidia/bf16', None , 'azure'
AI_BASE_URL = 'https://openrouter.ai/api/v1'  # Provider endpoint
# AI_ALLOW_FALLBACKS é determinado automaticamente: True se AI_PROVIDER é None, False caso contrário

# Model Selection - Change this to switch models
AI_MODEL = 'google/gemini-2.5-flash-lite-preview-09-2025'  # Default model
'''
    Models for OpenRouter:
    
        arcee-ai/trinity-mini:free       # Exceeded max tokens frequently
        nvidia/nemotron-nano-9b-v2:free
        nvidia/nemotron-3-nano-30b-a3b:free
        google/gemini-2.5-flash-lite-preview-09-2025
        openai/gpt-5-nano

'''
# AI Parameters
AI_TEMPERATURE = 0.1      # Low for consistent extraction (0.0-1.0)
AI_MAX_TOKENS = 8000      # Maximum tokens in response (sufficient for structured data extraction)
MAX_RETRIES = 2           # Retry attempts for API calls
MAX_PDF_TEXT_LENGTH = 20000  # Limit PDF text to prevent token overflow

# ============================================================================
# RESOLUTION EXTRACTION SCHEMA
# ============================================================================

RESOLUTION_FIELDS = [
    'numero_resolucao',
    'relacionada',
    'objeto',
    'data_inicial',
    'prazo_execucao',
    'vedado_utilizacao',
    'dotacao_orcamentaria'
]

# Excel column headers (Portuguese)
EXCEL_COLUMNS = [
    'Número da Resolução',
    'Relacionada',
    'Objeto',
    'Data Inicial',
    'Prazo Execução',
    'Vedado à Utilização',
    'Dotação Orçamentária',
    'Abreviação',
    'Link'
]

# Budget allocation to category mapping
BUDGET_CATEGORIES = {
    '301': 'Atenção Primária',
    '302': 'MAC',
    '303': 'Assistência Farmacêutica',
    '304': 'Vigilância Sanitária',
    '305': 'Vigilância Epidemiológica',
    '306': 'Alimentação e Nutrição',
    '122': 'ADM',
    '242': 'Assist. ao Portador de Deficiência'
}

# ============================================================================
# ENVIRONMENT CONFIGURATION
# ============================================================================

# Load API key from .env file (with UTF-8 BOM handling)
env_path = Path(__file__).parent.parent.parent / 'config' / '.env'
load_dotenv(env_path, encoding='utf-8-sig')  # utf-8-sig remove BOM automaticamente

# Get API key (only thing from .env)
OPENAI_API_KEY = os.getenv('OPENAI_API_KEY', '')

# ============================================================================
# LOGGING - Usando print() simples para mensagens limpas
# ============================================================================

# ============================================================================
# MAIN CLASS
# ============================================================================

class PDFToTableConverter:
    """
    AI-powered PDF to Excel converter for Brazilian government resolutions.

    Processes directory of resolution PDFs, extracts structured data using AI,
    and generates a consolidated Excel file with all resolutions as rows.
    """

    def __init__(self, api_key: Optional[str] = None, model: Optional[str] = None):
        """
        Initialize converter with OpenAI client.

        Args:
            api_key: OpenAI API key (optional, will use .env if not provided)
            model: AI model to use (optional, will use AI_MODEL constant if not provided)
        """
        self.api_key = api_key or OPENAI_API_KEY
        self.model = model or AI_MODEL
        self.max_retries = MAX_RETRIES
        self.enabled = False
        self.client = None

        if not self.api_key:
            print("⚠ API key não configurada - processamento de PDF desabilitado")
            print("  Configure OPENAI_API_KEY em src/config/.env")
            return

        try:
            # Configure client based on provider
            if AI_PROVIDER == 'openrouter' or (AI_BASE_URL and 'openrouter' in AI_BASE_URL.lower()):
                # OpenRouter com headers de atribuição
                self.client = OpenAI(
                    api_key=self.api_key,
                    base_url=AI_BASE_URL,
                    default_headers={
                        "HTTP-Referer": "https://github.com/marcomprado/bot_bb_daf",
                        "X-Title": "BOT bb"
                    }
                )
            elif AI_BASE_URL:
                # Outro provider personalizado
                self.client = OpenAI(
                    api_key=self.api_key,
                    base_url=AI_BASE_URL
                )
            else:
                # OpenAI direto
                self.client = OpenAI(api_key=self.api_key)

            self.enabled = True

        except Exception as e:
            print(f"✗ Erro ao inicializar IA: {e}")
            self.client = None
            self.enabled = False

    def _get_provider_config(self) -> Dict[str, Any]:
        """
        Retorna configuração de provider para OpenRouter via extra_body.
        """
        # Se não estiver usando OpenRouter, não precisa de provider config
        if not AI_BASE_URL or 'openrouter' not in AI_BASE_URL.lower():
            return {}

        # Se AI_PROVIDER é None, permite fallbacks (OpenRouter escolhe)
        if AI_PROVIDER is None:
            return {
                "extra_body": {
                    "provider": {"allow_fallbacks": True}
                }
            }

        # Se AI_PROVIDER é 'openai', não precisa de provider config
        if AI_PROVIDER == 'openai':
            return {}

        # Configuração de provider para OpenRouter com provider específico
        provider_config = {
            "order": [AI_PROVIDER],  # Lista de providers em ordem de preferência
            "allow_fallbacks": False  # Não permite outros providers
        }

        return {
            "extra_body": {
                "provider": provider_config
            }
        }

    def process_file_list(self, pdf_files: List[Dict], output_dir: str) -> Dict:
        """
        Process a specific list of PDF files (from bot run).

        Args:
            pdf_files: List of dicts with 'caminho' (path), optional 'url'
            output_dir: Directory to save Excel file

        Returns:
            Dict with success status, Excel path, statistics, errors
        """
        if not self.enabled:
            return {
                'success': False,
                'error': 'AI client disabled - check OPENAI_API_KEY in .env'
            }

        results = []
        successful = 0
        failed = 0

        print(f"\n→ Processando {len(pdf_files)} PDFs com IA...\n")

        for idx, pdf_info in enumerate(pdf_files, 1):
            pdf_path = pdf_info.get('caminho')
            file_link = pdf_info.get('url', 'NÃO INFORMADO')

            print(f"  [{idx}/{len(pdf_files)}] {Path(pdf_path).name}")

            result = self.process_single_pdf(pdf_path, file_link)
            results.append(result)

            if result.get('success', False):
                successful += 1
            else:
                failed += 1

        # Generate Excel with ALL results (success + failures)
        date_str = datetime.now().strftime('%Y-%m-%d')
        excel_filename = f'resolucoes-{date_str}.xlsx'
        excel_path = os.path.join(output_dir, excel_filename)

        try:
            self.generate_excel(results, excel_path)

            # Cleanup PDFs and folder after successful Excel generation
            cleanup_result = self._cleanup_successful_pdfs(results, output_dir, excel_path)

            if cleanup_result['pdfs_deleted'] > 0:
                print(f"\n  🗑️  {cleanup_result['pdfs_deleted']} PDFs processados foram removidos")

            # Use new Excel path if folder was deleted and Excel moved
            final_excel_path = cleanup_result.get('new_excel_path', excel_path)

            return {
                'success': True,
                'excel_path': final_excel_path,
                'total_processed': len(results),
                'successful': successful,
                'failed': failed,
                'pdfs_deleted': cleanup_result['pdfs_deleted'],
                'folder_deleted': cleanup_result.get('folder_deleted', False),
                'results': results
            }
        except Exception as e:
            print(f"✗ Erro ao gerar Excel: {e}")
            return {
                'success': False,
                'error': f'Excel generation failed: {e}',
                'total_processed': len(results),
                'successful': successful,
                'failed': failed
            }

    def process_single_pdf(self, pdf_path: str, file_link: Optional[str] = None) -> Dict:
        """
        Process single PDF - returns success dict or error dict.

        Args:
            pdf_path: Path to PDF file
            file_link: Optional URL link to PDF source

        Returns:
            Dict with success status and extracted_data
        """
        filename = Path(pdf_path).name
        timestamp = datetime.now().strftime('%d/%m/%Y %H:%M:%S')

        # Base result structure
        result = {
            'success': False,
            'file_path': pdf_path,
            'file_name': filename,
            'file_link': file_link or 'NÃO INFORMADO'
        }

        try:
            # Validate file exists
            if not Path(pdf_path).exists():
                raise FileNotFoundError(f"File not found: {pdf_path}")

            # Extract text
            pdf_text = self._extract_text_from_pdf(pdf_path)
            if not pdf_text or len(pdf_text.strip()) < 100:
                raise ValueError(f"Insufficient text extracted: {len(pdf_text)} chars")

            # AI extraction
            extracted_data = self._extract_resolution_data(pdf_text)
            if 'error' in extracted_data:
                raise ValueError(f"AI extraction failed: {extracted_data['error']}")

            # Add metadata and categorization
            extracted_data['arquivo'] = filename
            extracted_data['link'] = file_link or 'NÃO INFORMADO'
            extracted_data['data_processamento'] = timestamp
            extracted_data['abreviacao'] = self._categorize_by_budget_allocation(
                extracted_data.get('dotacao_orcamentaria', '')
            )
            extracted_data['status'] = 'SUCESSO'

            result['success'] = True
            result['extracted_data'] = extracted_data
            result['pdf_path'] = pdf_path
            return result

        except Exception as e:
            # Create error row for Excel
            print(f"    ✗ Erro: {str(e)[:80]}")

            error_data = {
                'numero_resolucao': 'ERRO',
                'relacionada': 'ERRO',
                'objeto': f'Falha ao processar: {str(e)}',
                'data_inicial': 'ERRO',
                'prazo_execucao': 'ERRO',
                'vedado_utilizacao': 'ERRO',
                'dotacao_orcamentaria': 'ERRO',
                'abreviacao': 'ERRO',
                'link': file_link or 'NÃO INFORMADO',
                'arquivo': filename,
                'data_processamento': timestamp,
                'status': f'FALHA: {str(e)[:100]}'  # Truncate long errors
            }

            result['extracted_data'] = error_data
            result['error'] = str(e)
            result['pdf_path'] = pdf_path
            return result

    def generate_excel(self, results: List[Dict], output_path: str) -> str:
        """
        Generate formatted Excel with success and error rows.

        Args:
            results: List of processing results
            output_path: Path to save Excel file

        Returns:
            Path to saved Excel file
        """
        # Create workbook
        wb = Workbook()
        ws = wb.active
        ws.title = "Resoluções"

        # Define styles
        header_font = Font(name='Arial', size=11, bold=True, color='000000')
        header_fill = PatternFill(start_color='D3D3D3', end_color='D3D3D3', fill_type='solid')
        data_font = Font(name='Arial', size=10, color='000000')
        error_font = Font(name='Arial', size=10, color='FF0000', bold=True)
        error_value_font = Font(name='Arial', size=10, color='FF0000')

        alignment_center = Alignment(horizontal='center', vertical='center')
        alignment_left = Alignment(horizontal='left', vertical='center', wrap_text=True)

        border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )

        # Write header row
        for col_idx, header in enumerate(EXCEL_COLUMNS, 1):
            cell = ws.cell(row=1, column=col_idx, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = alignment_center
            cell.border = border

        # Write data rows
        for row_idx, result in enumerate(results, 2):
            data = result.get('extracted_data', {})
            is_error = not result.get('success', False)

            # Determine font based on success/failure
            row_font = error_value_font if is_error else data_font

            # Column data
            row_data = [
                data.get('numero_resolucao', 'NÃO INFORMADO'),
                data.get('relacionada', 'NÃO INFORMADO'),
                data.get('objeto', 'NÃO INFORMADO'),
                data.get('data_inicial', 'NÃO INFORMADO'),
                data.get('prazo_execucao', 'NÃO INFORMADO'),
                data.get('vedado_utilizacao', 'NÃO INFORMADO'),
                data.get('dotacao_orcamentaria', 'NÃO INFORMADO'),
                data.get('abreviacao', 'NÃO CLASSIFICADO'),
                data.get('link', 'NÃO INFORMADO')
            ]

            for col_idx, value in enumerate(row_data, 1):
                cell = ws.cell(row=row_idx, column=col_idx, value=value)

                # Apply error styling for ERRO values
                if value == 'ERRO' or (isinstance(value, str) and value.startswith('FALHA:')):
                    cell.font = error_font
                else:
                    cell.font = row_font

                # Alignment
                if col_idx in [1, 4, 5]:  # Dates, number
                    cell.alignment = alignment_center
                else:
                    cell.alignment = alignment_left

                cell.border = border

        # Set column widths
        column_widths = [20, 20, 60, 15, 15, 50, 25, 18, 50]
        for col_idx, width in enumerate(column_widths, 1):
            ws.column_dimensions[chr(64 + col_idx)].width = width

        # Freeze header row
        ws.freeze_panes = 'A2'

        # Save
        wb.save(output_path)

        return output_path

    def _cleanup_successful_pdfs(self, results: List[Dict], output_dir: str, excel_path: str) -> Dict:
        """
        Remove PDF files that were successfully processed.
        Keeps failed PDFs for manual review.
        Checks each unique directory containing PDFs and deletes empty month folders.
        """
        import os

        deleted_pdfs = 0
        folders_deleted = []

        # Step 1: Delete successful PDFs and track their directories
        pdf_directories = set()

        for result in results:
            if result.get('success', False):
                pdf_path = result.get('pdf_path')

                if pdf_path:
                    try:
                        if os.path.exists(pdf_path):
                            pdf_dir = os.path.dirname(pdf_path)
                            pdf_directories.add(pdf_dir)

                            os.remove(pdf_path)
                            deleted_pdfs += 1
                    except Exception as e:
                        pdf_filename = os.path.basename(pdf_path)
                        print(f"    ⚠ Não foi possível apagar {pdf_filename}: {e}")

        # Step 2: Check each directory and delete if empty (month folders)
        for pdf_dir in pdf_directories:
            try:
                # Only consider directories that are NOT the Excel output directory
                if pdf_dir == output_dir:
                    continue

                remaining_files = [f for f in os.listdir(pdf_dir)
                                 if os.path.isfile(os.path.join(pdf_dir, f))]

                if len(remaining_files) == 0:
                    # Folder is empty - delete it
                    os.rmdir(pdf_dir)
                    folders_deleted.append(pdf_dir)
                    print(f"  📁 Pasta vazia removida: {os.path.basename(pdf_dir)}")

            except Exception as e:
                print(f"    ⚠ Erro ao limpar pasta {pdf_dir}: {e}")

        return {
            'pdfs_deleted': deleted_pdfs,
            'folders_deleted': folders_deleted,
            'total_folders_deleted': len(folders_deleted)
        }

    # ========================================================================
    # PRIVATE METHODS - PDF Processing
    # ========================================================================

    def _extract_text_from_pdf(self, pdf_path: str) -> str:
        """
        Extract text content from PDF using pymupdf4llm.
        """
        try:
            # Validate file exists and has reasonable size
            pdf_file = Path(pdf_path)
            if not pdf_file.exists():
                raise FileNotFoundError(f"PDF file does not exist: {pdf_path}")

            file_size = pdf_file.stat().st_size
            if file_size == 0:
                raise ValueError(f"PDF file is empty: {pdf_path}")

            # Extract text using pymupdf4llm
            text = pymupdf4llm.to_markdown(str(pdf_path))

            if text:
                return text
            else:
                return ""

        except (FileNotFoundError, PermissionError, ValueError, Exception):
            return ""

    def _extract_resolution_data(self, pdf_text: str) -> Dict[str, Any]:
        """
        Extract structured resolution data using AI with specific prompt.
        """
        try:
            # Create the specialized prompt for resolution extraction
            system_prompt = self._get_resolution_extraction_prompt()

            # Prepare user content with PDF text (limit to prevent token overflow)
            limited_text = pdf_text[:MAX_PDF_TEXT_LENGTH]
            user_content = f"""Analise o seguinte texto extraído de um PDF de resolução e extraia os dados estruturados conforme solicitado:

TEXTO DO PDF:
{limited_text}

Proceda com a análise e retorne os dados no formato JSON especificado."""

            # Create messages for API call
            messages = [
                {"role": "system", "content": system_prompt},
                {"role": "user", "content": user_content}
            ]

            # Call OpenAI API
            response = self._chat_completion(messages, AI_TEMPERATURE)

            # Parse the JSON response (extract from markdown if needed)
            content = response['content'].strip()
            json_content = self._extract_json_from_response(content)

            extracted_data = json.loads(json_content)

            # Add API usage info
            extracted_data['_ai_metadata'] = {
                'tokens_used': response['usage']['total_tokens'],
                'model': response['model']
            }

            return extracted_data

        except json.JSONDecodeError as e:
            print(f"    ✗ Erro ao processar resposta da IA")
            raw_content = response.get('content', '') if 'response' in locals() else ''
            return {
                'error': 'Failed to parse AI response',
                'raw_content': raw_content
            }

        except Exception as e:
            return {'error': str(e)}

    def _chat_completion(self, messages: List[Dict[str, str]], temperature: float = 0.1) -> Dict[str, Any]:
        """
        Call OpenAI API with retry logic.

        """
        for attempt in range(self.max_retries):
            try:
                if not self.client:
                    raise ValueError("OpenAI client not initialized")

                # Preparar configuração base
                request_params = {
                    "model": self.model,
                    "messages": messages,
                    "temperature": temperature,
                    "max_tokens": AI_MAX_TOKENS
                }

                # Adicionar provider config se necessário (para OpenRouter)
                provider_config = self._get_provider_config()
                request_params.update(provider_config)

                # Fazer requisição com ou sem provider config
                response = self.client.chat.completions.create(**request_params)

                result = {
                    'content': response.choices[0].message.content,
                    'usage': {
                        'prompt_tokens': response.usage.prompt_tokens,
                        'completion_tokens': response.usage.completion_tokens,
                        'total_tokens': response.usage.total_tokens
                    },
                    'model': response.model,
                    'finish_reason': response.choices[0].finish_reason
                }

                return result

            except openai.RateLimitError as e:
                print(f"    ⏳ Rate limit - tentativa {attempt+1}/{self.max_retries}")
                if attempt < self.max_retries - 1:
                    time.sleep(2 ** attempt)  # Exponential backoff
                else:
                    raise ValueError(f"Rate limit exceeded after {self.max_retries} attempts")

            except openai.AuthenticationError as e:
                raise ValueError(f"Invalid API key: {e}")

            except Exception as e:
                print(f"    ✗ Erro API (tentativa {attempt+1}/{self.max_retries})")
                if attempt < self.max_retries - 1:
                    time.sleep(2 ** attempt)
                else:
                    raise ValueError(f"API error after {self.max_retries} attempts: {e}")

        raise ValueError("Max retries exceeded for chat completion")

    def _extract_json_from_response(self, content: str) -> str:
        """
        Extract JSON from AI response, handling markdown code blocks.
        """
        content = content.strip()

        # Handle markdown-wrapped JSON response
        if content.startswith('```json'):
            # Extract JSON from markdown code block
            start_idx = content.find('{')
            end_idx = content.rfind('}') + 1
            if start_idx != -1 and end_idx != 0:
                return content[start_idx:end_idx]

        return content

    def _get_resolution_extraction_prompt(self) -> str:
        """
            System prompt for AI extraction
        """
        return """Você é um assistente especializado em análise de documentos legais. Sua tarefa é extrair informações específicas de um PDF e retornar os dados em formato estruturado para composição de uma tabela.

INSTRUÇÕES GERAIS:
• Analise cuidadosamente todo o conteúdo do PDF fornecido
• Extraia apenas as informações solicitadas
• Seja preciso nas extrações, mantendo a fidelidade ao texto original
• Se alguma informação não estiver presente, retorne "NÃO INFORMADO"
• !!! Se atente para que a resposta final esteja estritamente no formato JSON especificado e nenhum texto adicional seja incluído.

DADOS A EXTRAIR:

1. NÚMERO DA RESOLUÇÃO
• Formato esperado: "xxxxx/20XX"
• Localização: Sempre no cabeçalho, início do documento ou no titulo.
• Exemplo: "12345/2023"
• Retornar: O número da resolução (no formato xxxxx/20XX).

2. RELACIONADA
• Descrição: Verificar se a resolução cita, modifica, altera ou revoga outra resolução
• Palavras-chave: "Altera a Resolução", "modifica", "altera", "revoga", "em substituição", "complementa" .
• Retornar: O número da resolução relacionada (mesmo formato xxxxx/20XX) ou "NÃO INFORMADO"

3. OBJETO
• Descrição: Extrair integralmente do primeiro parágrafo da resolução
• Características: Geralmente começa após o número e data, é a descrição do propósito da resolução.
• Retornar: Texto baseado no primeiro parágrafo, sem omissões de ideias, mas resumido.

4. DATA INICIAL
• Formato esperado: "DD/MM/AAAA"
• Localização: Data que aparece logo após o número da resolução
• Retornar no formato esperado "DD/MM/AAAA". -- Exemplo: "15/03/2023"

5. PRAZO EXECUÇÃO
• Descrição: Data estimada mencionada no texto para execução/vigência
• Cálculo: Se expresso em meses/anos a partir da data inicial:
  - Cada mês = 30 dias
  - Cada ano = 365 dias
• Formato de retorno: "DD/MM/AAAA"
• Se não especificado: "NÃO INFORMADO"

6. VEDADO A UTILIZAÇÃO
• Descrição: Parágrafo ou trecho que detalha restrições de uso de recursos/verbas
• Palavras-chave: "vedado", "proibido", "não poderá ser utilizado", "fica vedada"
• Retornar: Texto completo do parágrafo que contém as vedações.

7. DOTAÇÃO ORÇAMENTÁRIA
• Descrição: Conjunto numérico que segue imediatamente após a expressão "dotação orçamentária"
• Formato: Sequência de números, pontos e traços (ex: "12.345.67.89.123")
• Atenção especial: Procurar pelos códigos 301, 302, 303, 304, 305, 306, 122, 242 dentro da dotação
• Retornar: Toda a sequência numérica da dotação orçamentária completa.

OBSERVAÇÕES IMPORTANTES:
• Mantenha fidelidade absoluta ao texto original
• Não interprete ou parafraseie as informações
• Em caso de dúvida sobre localização de dados, busque por padrões similares
• Datas devem estar sempre no formato DD/MM/AAAA
• Para campos não encontrados, use exatamente "NÃO INFORMADO"

Proceda com a análise do PDF fornecido e retorne os dados no formato especificado.

!!! FORMATO DE RESPOSTA FINAL:
Retorne os dados extraídos dentro do seguinte formato JSON (Se Mantenha EXTREMAMENTE fiel a esse formato) :
{
  "numero_resolucao": "Resposta_aqui",
  "relacionada": "Resposta_aqui",
  "objeto": "Resposta_aqui",
  "data_inicial": "Resposta_aqui",
  "prazo_execucao": "Resposta_aqui",
  "vedado_utilizacao": "Resposta_aqui",
  "dotacao_orcamentaria": "Resposta_aqui"
}

"""

    # ========================================================================
    # PRIVATE METHODS - Validation and Categorization
    # ========================================================================

    def _validate_resolution_number(self, number: str) -> bool:
        """
        Validate resolution number format (xxxxx/20XX).
        """
        if not number or number == "NÃO INFORMADO":
            return True
        pattern = r'^\d{4,5}/20\d{2}$'
        return bool(re.match(pattern, number))

    def _validate_date(self, date_str: str) -> bool:
        """
        Validate Brazilian date format (DD/MM/AAAA).
        """
        if not date_str or date_str == "NÃO INFORMADO":
            return True
        pattern = r'^\d{2}/\d{2}/\d{4}$'
        if not re.match(pattern, date_str):
            return False

        # Optional: validate actual date
        try:
            datetime.strptime(date_str, '%d/%m/%Y')
            return True
        except ValueError:
            return False

    def _categorize_by_budget_allocation(self, dotacao: str) -> str:
        """
        Map budget allocation codes to category abbreviations.
        """
        if not dotacao or dotacao == "NÃO INFORMADO":
            return "NÃO CLASSIFICADO"

        for code, category in BUDGET_CATEGORIES.items():
            if code in dotacao:
                return category

        return "NÃO CLASSIFICADO"
